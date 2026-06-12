"""
XLSX Builder for BI Dashboard Export.

Creates an 8-sheet Excel workbook from BI backend API data.
Each sheet corresponds to one BI tab with cards, charts, and tables.
"""

import io
import logging
import math
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from lxml import etree

import openpyxl
import openpyxl.chart.marker
from openpyxl.chart import (
    BarChart, LineChart, PieChart, DoughnutChart, RadarChart,
    Reference, Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XlImage
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from xlsx_theme import (
    XlsxStyles, ROW_HEIGHTS, COL_WIDTHS, TAB_COLORS,
    CHART_COLORS_HEX, CHART_WIDTH_CM, CHART_HEIGHT_CM,
    CHART_SMALL_W, CHART_SMALL_H, CHART_ROWS_STANDARD, CHART_ROWS_SMALL,
    PRIMARY, SEMANTIC,
)
from bi_client import _localize

logger = logging.getLogger(__name__)


# ── Sheet names ──────────────────────────────────────────────────────────────

SHEET_NAMES = {
    "zh": [
        "执行摘要", "SLA分析", "事件分析", "变更分析",
        "请求分析", "问题分析", "跨流程关联", "Workforce",
    ],
    "en": [
        "Executive Summary", "SLA Analysis", "Incident Analysis", "Change Analysis",
        "Request Analysis", "Problem Analysis", "Cross-Process", "Workforce",
    ],
}

TAB_ID_TO_SHEET_INDEX = {
    "executive-summary": 0,
    "sla-analysis": 1,
    "incident-analysis": 2,
    "change-analysis": 3,
    "request-analysis": 4,
    "problem-analysis": 5,
    "cross-process": 6,
    "workforce": 7,
}


def _t(text_en: str, text_zh: str, language: str) -> str:
    """Return localized string."""
    return text_zh if language == "zh" else text_en


# =============================================================================
# Chart Renderer — converts BI chart data to openpyxl chart objects
# =============================================================================

class BiChartRenderer:
    """Renders BI ChartSection data into native openpyxl chart objects."""

    def __init__(self, wb: openpyxl.Workbook, language: str = "zh"):
        self.wb = wb
        self.language = language
        self._data_ws = wb.create_sheet("_ChartData")
        self._data_ws.sheet_state = "hidden"
        self._data_row = 1

    def _write_data(self, headers: list, rows: list) -> Tuple[str, int, int]:
        """Write data block to hidden sheet. Returns (sheet_title, start_row, end_row)."""
        start = self._data_row
        for c, h in enumerate(headers, 1):
            self._data_ws.cell(row=start, column=c, value=h)
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row, 1):
                self._data_ws.cell(row=start + 1 + r_idx, column=c_idx, value=val)
        end = start + len(rows)
        self._data_row = end + 2
        return self._data_ws.title, start, end

    def _color(self, idx: int) -> str:
        return CHART_COLORS_HEX[idx % len(CHART_COLORS_HEX)]

    # Font for chart text — matches BI frontend
    CHART_FONT = "Microsoft YaHei"
    CHART_FONT_SIZE = 900  # hundredths of a point → 9pt

    def _chart_font_richtext(self, color: str = None):
        """Create a RichText with the standard chart font. Optional color for axis tick labels."""
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
        font = DrawingFont(typeface=self.CHART_FONT)
        if color:
            cp = CharacterProperties(latin=font, ea=font, sz=self.CHART_FONT_SIZE, solidFill=color)
        else:
            cp = CharacterProperties(latin=font, ea=font, sz=self.CHART_FONT_SIZE)
        pp = ParagraphProperties(defRPr=cp)
        return RichText(p=[Paragraph(pPr=pp)])

    def _apply_style(self, chart, title: str, width_cm: float = CHART_WIDTH_CM, height_cm: float = CHART_HEIGHT_CM,
                     left_axis_color: str = None, right_axis_color: str = None):
        chart.width = width_cm
        chart.height = height_cm
        chart.style = None
        chart.title = None

        # ── Font for all chart text ──
        font_rt = self._chart_font_richtext()
        left_y_font = self._chart_font_richtext(color=left_axis_color) if left_axis_color else font_rt
        right_y_font = self._chart_font_richtext(color=right_axis_color) if right_axis_color else font_rt

        # ── Legend ──
        if chart.legend is not None:
            chart.legend.position = "b"
            chart.legend.overlay = False
            chart.legend.txPr = font_rt

        # ── Axis tick label visibility (bar / line / combo charts only) ──
        has_axes = hasattr(chart, 'x_axis') and hasattr(chart, 'y_axis')
        is_radar = isinstance(chart, RadarChart)

        if has_axes:
            from openpyxl.chart.shapes import GraphicalProperties as GP

            # Primary X-axis
            if chart.x_axis is not None:
                chart.x_axis.delete = False
                chart.x_axis.tickLblPos = "low"
                chart.x_axis.majorTickMark = "none"
                chart.x_axis.minorTickMark = "none"
                chart.x_axis.txPr = font_rt
                # Light axis line
                chart.x_axis.graphicalProperties = GP()
                chart.x_axis.graphicalProperties.line.solidFill = "D0D5DD"
                chart.x_axis.graphicalProperties.line.width = 8000

            # Primary Y-axis (left) — colored if requested
            if chart.y_axis is not None:
                # Radar charts use radial grid lines; hide the linear Y-axis
                chart.y_axis.delete = True if is_radar else False
                chart.y_axis.tickLblPos = "nextTo"
                chart.y_axis.majorTickMark = "none"
                chart.y_axis.minorTickMark = "none"
                chart.y_axis.txPr = left_y_font
                if not chart.y_axis.numFmt or chart.y_axis.numFmt == 'General':
                    chart.y_axis.numFmt = '#,##0'
                # Subtle grid lines (skip for radar — it has its own polar grid)
                if not is_radar:
                    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
                    chart.y_axis.majorGridlines.graphicalProperties = GP()
                    chart.y_axis.majorGridlines.graphicalProperties.line.solidFill = "E8ECF0"
                    chart.y_axis.majorGridlines.graphicalProperties.line.width = 5000
                    chart.y_axis.majorGridlines.graphicalProperties.line.dashStyle = "solid"
                # No axis line on Y
                chart.y_axis.graphicalProperties = GP()
                chart.y_axis.graphicalProperties.line.noFill = True

            # For combo charts: configure secondary Y-axis on child line charts
            # Note: _charts[0] is the bar chart itself (shares y_axis with parent),
            # _charts[1] is the line chart (has its own secondary y_axis).
            children = getattr(chart, '_charts', [])
            for child in children:
                if hasattr(child, 'y_axis') and child.y_axis is not None and child.y_axis is not chart.y_axis:
                    child.y_axis.delete = False
                    child.y_axis.tickLblPos = "nextTo"
                    child.y_axis.majorTickMark = "none"
                    child.y_axis.minorTickMark = "none"
                    child.y_axis.txPr = right_y_font
                    if not child.y_axis.numFmt or child.y_axis.numFmt == 'General':
                        child.y_axis.numFmt = '#,##0'
                    child.y_axis.graphicalProperties = GP()
                    child.y_axis.graphicalProperties.line.noFill = True

            # ── Plot area styling (only for charts with axes) ──
            chart.plot_area.graphicalProperties = GP()
            chart.plot_area.graphicalProperties.noFill = True
            chart.plot_area.graphicalProperties.line.noFill = True

    def _parse_pipe_data(self, items: List[Dict]) -> Tuple[List[str], List[str], List[List[float]]]:
        """Parse pipe-delimited ChartDatum labels into categories and series values.

        Returns: (categories, series_names, data_matrix)
        """
        if not items:
            return [], [], []

        first_label = str(items[0].get("label", ""))
        parts = first_label.split("|")

        categories = []
        series_count = len(parts) - 1
        data_matrix = [[] for _ in range(series_count)]

        for item in items:
            label = str(item.get("label", ""))
            parts = label.split("|")
            categories.append(parts[0] if parts else label)
            for i in range(series_count):
                val = 0.0
                if i + 1 < len(parts):
                    try:
                        val = float(parts[i + 1])
                    except (ValueError, TypeError):
                        pass
                data_matrix[i].append(val)

        # Default series names
        series_names = [f"Series {i + 1}" for i in range(series_count)]
        return categories, series_names, data_matrix

    def render(self, chart_data: Dict) -> Optional[Any]:
        """Dispatch to appropriate renderer based on chart type."""
        chart_type = chart_data.get("type", "bar")
        method_name = f"render_{chart_type.replace('-', '_')}"
        renderer = getattr(self, method_name, self.render_bar)
        try:
            return renderer(chart_data)
        except Exception as e:
            logger.warning("Chart render failed for type=%s title=%s: %s", chart_type, chart_data.get("title"), e)
            return None

    def render_pie(self, chart_data: Dict) -> Optional[PieChart]:
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        if not items:
            return None

        labels = [str(i.get("label", f"Item {idx}")) for idx, i in enumerate(items)]
        values = [float(v) if (v := i.get("value", 0)) is not None else 0.0 for i in items]
        if sum(values) == 0:
            return None

        sheet_title, start, end = self._write_data(
            [_t("Category", "类别", self.language), _t("Value", "数值", self.language)],
            [[l, v] for l, v in zip(labels, values)]
        )

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        # Hide series title from legend and labels
        chart.series[0].title = None

        colors = config.get("colors", []) if config else []
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = "FFFFFF"
            pt.graphicalProperties.line.width = 12000
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, chart_data.get("title", "Pie Chart"))
        # Data labels: show only category name + percentage, no "Value" / series name
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLeaderLines = True
        return chart

    def render_line(self, chart_data: Dict) -> Optional[LineChart]:
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        if not items:
            return None

        categories, series_names, data_matrix = self._parse_pipe_data(items)
        if config and config.get("series"):
            series_names = config["series"]

        if not categories:
            return None

        # Build headers and rows
        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for i, cat in enumerate(categories):
            row = [cat] + [data_matrix[j][i] if j < len(data_matrix) else 0 for j in range(len(series_names))]
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)

        chart = LineChart()
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx in range(2, 2 + len(series_names)):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Apply colors, disable smooth, add markers
        colors = config.get("colors", []) if config else []
        for i, series in enumerate(chart.series):
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 22000  # ~1.75pt
            series.smooth = False  # straight lines, not curves
            # Add data point markers (circles with fill)
            marker = openpyxl.chart.marker.Marker(symbol="circle", size=5)
            marker.graphicalProperties.solidFill = color
            marker.graphicalProperties.line.noFill = True
            series.marker = marker

        self._apply_style(chart, chart_data.get("title", "Line Chart"))
        # No axis titles — frontend uses only bottom legend, not axis text labels
        if not chart.y_axis.numFmt or chart.y_axis.numFmt == 'General':
            chart.y_axis.numFmt = '#,##0'
        return chart

    def render_bar(self, chart_data: Dict) -> Optional[BarChart]:
        """Horizontal bar chart (barDir='bar')."""
        items = chart_data.get("items", [])
        if not items:
            return None

        labels = [str(i.get("label", f"Item {idx}")) for idx, i in enumerate(items)]
        values = [float(v) if (v := i.get("value", 0)) is not None else 0.0 for i in items]
        if sum(values) == 0:
            return None

        sheet_title, start, end = self._write_data(
            [_t("Category", "类别", self.language), _t("Value", "数值", self.language)],
            [[l, v] for l, v in zip(labels, values)]
        )

        chart = BarChart()
        chart.type = "bar"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            pt.graphicalProperties.line.noFill = True
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, chart_data.get("title", "Bar Chart"))
        chart.legend = None
        chart.x_axis.numFmt = '#,##0'
        chart.y_axis.numFmt = '#,##0'
        return chart

    def render_column(self, chart_data: Dict) -> Optional[BarChart]:
        """Vertical column chart (barDir='col')."""
        items = chart_data.get("items", [])
        if not items:
            return None

        labels = [str(i.get("label", f"Item {idx}")) for idx, i in enumerate(items)]
        values = [float(v) if (v := i.get("value", 0)) is not None else 0.0 for i in items]
        if sum(values) == 0:
            return None

        sheet_title, start, end = self._write_data(
            [_t("Category", "类别", self.language), _t("Value", "数值", self.language)],
            [[l, v] for l, v in zip(labels, values)]
        )

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        config = chart_data.get("config", {})
        colors = config.get("colors", []) if config else []
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.noFill = True
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, chart_data.get("title", "Column Chart"))
        chart.legend = None
        chart.y_axis.numFmt = '#,##0'
        return chart

    def render_grouped_bar(self, chart_data: Dict) -> Optional[BarChart]:
        """Grouped bar chart (clustered)."""
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        if not items:
            return None

        categories, series_names, data_matrix = self._parse_pipe_data(items)
        if config and config.get("series"):
            series_names = config["series"]

        if not categories:
            return None

        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for i, cat in enumerate(categories):
            row = [cat] + [data_matrix[j][i] if j < len(data_matrix) else 0 for j in range(len(series_names))]
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx in range(2, 2 + len(series_names)):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        colors = config.get("colors", []) if config else []
        for i, series in enumerate(chart.series):
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.noFill = True

        self._apply_style(chart, chart_data.get("title", "Grouped Bar Chart"))
        chart.y_axis.numFmt = '#,##0'
        return chart

    def render_stacked_bar(self, chart_data: Dict) -> Optional[BarChart]:
        """Stacked bar chart."""
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        if not items:
            return None

        categories, series_names, data_matrix = self._parse_pipe_data(items)
        if config and config.get("series"):
            series_names = config["series"]

        if not categories:
            return None

        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for i, cat in enumerate(categories):
            row = [cat] + [data_matrix[j][i] if j < len(data_matrix) else 0 for j in range(len(series_names))]
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100  # Required for correct stacking in WPS / older Excel
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx in range(2, 2 + len(series_names)):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        colors = config.get("colors", []) if config else []
        for i, series in enumerate(chart.series):
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            series.graphicalProperties.solidFill = color
            # Thin white border between stacked segments
            series.graphicalProperties.line.solidFill = "FFFFFF"
            series.graphicalProperties.line.width = 6000

        self._apply_style(chart, chart_data.get("title", "Stacked Bar Chart"))
        chart.y_axis.numFmt = '#,##0'
        return chart

    # Charts whose right-axis series is a percentage (0-100)
    PERCENTAGE_RIGHT_AXIS_COMBOS = {
        "request-sla-time", "change-success-trend", "incident-volume-trend",
    }
    # Charts where both axes are percentages (0-100)
    PERCENTAGE_CHARTS = {
        "sla-trend", "req-sla-trend",
    }
    # Charts where right axis is satisfaction score (0-5, step 1)
    SATISFACTION_RIGHT_AXIS = {
        "request-volume-trend",   # 请求单量 + 平均满意度
        "req-sla-trend",          # SLA达成率 + 平均满意度
    }
    # Per-chart right axis config: (min, max, majorUnit, numFmt)
    RIGHT_AXIS_CONFIG = {
        "trend":                None,                    # auto — health score + event count
        "request-volume-trend": (0, 5, 1, '0.0'),      # 满意度 1-5
        "req-sla-trend":       (0, 5, 1, '0.0'),       # 满意度 1-5
        "request-sla-time":    (0, 100, 20, '0"%"'),    # SLA达成率 %
        "change-success-trend": (0, 100, 20, '0"%"'),   # 成功率 %
        "incident-volume-trend": (0, 100, 20, '0"%"'),  # SLA达成率 %
        "problem-volume-trend": None,                    # auto — both are counts
        "cross-change-incident-trend": None,             # auto — both are counts
        "cross-request-incident-overlap": None,          # auto — both are counts
    }

    def render_combo(self, chart_data: Dict) -> Optional[BarChart]:
        """Combo chart: bar (left Y-axis) + lines (shared right Y-axis)."""
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        chart_id = chart_data.get("id", "")
        if not items:
            return None

        categories, series_names, data_matrix = self._parse_pipe_data(items)
        if config and config.get("series"):
            series_names = config["series"]

        if not categories or len(series_names) < 2:
            return self.render_bar(chart_data)

        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for i, cat in enumerate(categories):
            row = [cat] + [data_matrix[j][i] if j < len(data_matrix) else 0 for j in range(len(series_names))]
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)
        colors = config.get("colors", []) if config else []

        # ── Bar chart: first series on LEFT Y-axis ──
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "clustered"
        count_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        bar.add_data(count_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.y_axis.numFmt = '#,##0'

        if chart_id in self.PERCENTAGE_CHARTS:
            bar.y_axis.scaling.min = 0
            bar.y_axis.scaling.max = 100
            bar.y_axis.numFmt = '0"%"'

        if colors:
            bar.series[0].graphicalProperties.solidFill = colors[0].lstrip("#")
            bar.series[0].graphicalProperties.line.noFill = True
        # Bar gap width
        bar.series[0].graphicalProperties.solidFill  # ensure set above

        # ── Line chart: ALL remaining series share ONE right Y-axis ──
        line = LineChart()
        for idx in range(1, len(series_names)):
            line_ref = Reference(self._data_ws, min_col=2 + idx, min_row=start, max_row=end)
            line.add_data(line_ref, titles_from_data=True)

        # Configure right Y-axis (shared by all line series)
        line.y_axis.axId = 200
        line.y_axis.numFmt = '#,##0'
        line.y_axis.crosses = "max"

        # Per-chart right axis scale configuration
        axis_cfg = self.RIGHT_AXIS_CONFIG.get(chart_id)
        if axis_cfg is not None:
            rmin, rmax, rstep, rfmt = axis_cfg
            line.y_axis.scaling.min = rmin
            line.y_axis.scaling.max = rmax
            line.y_axis.scaling.majorUnit = rstep
            line.y_axis.numFmt = rfmt
        else:
            # Auto-detect: check if any series name suggests percentage
            right_is_pct = (
                chart_id in self.PERCENTAGE_RIGHT_AXIS_COMBOS
                or chart_id in self.PERCENTAGE_CHARTS
            )
            if not right_is_pct:
                for idx in range(1, len(series_names)):
                    sname = series_names[idx] if idx < len(series_names) else ""
                    if "率" in sname or "%" in sname or "rate" in sname.lower():
                        right_is_pct = True
                        break
            if right_is_pct:
                line.y_axis.scaling.min = 0
                line.y_axis.scaling.max = 100
                line.y_axis.numFmt = '0"%"'

        # Style each line series: color, markers, no smooth
        for idx, series in enumerate(line.series):
            s_idx = idx + 1  # index into series_names/colors (0-based, offset by 1)
            color = colors[s_idx].lstrip("#") if s_idx < len(colors) else self._color(s_idx)
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 22000  # ~1.75pt
            series.smooth = False
            series.marker = openpyxl.chart.marker.Marker(symbol="circle", size=5)
            series.marker.graphicalProperties.solidFill = color

        bar += line

        # Pass axis colors to _apply_style so Y-axis tick labels match series colors
        left_color = colors[0].lstrip("#") if colors else None
        right_color = colors[1].lstrip("#") if len(colors) > 1 else None

        self._apply_style(bar, chart_data.get("title", "Combo Chart"),
                          left_axis_color=left_color, right_axis_color=right_color)
        return bar

    def render_combo_line(self, chart_data: Dict) -> Optional[LineChart]:
        """Dual-line combo: first series as line (left Y) + second series as line (right Y)."""
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        chart_id = chart_data.get("id", "")
        if not items:
            return None

        categories, series_names, data_matrix = self._parse_pipe_data(items)
        if config and config.get("series"):
            series_names = config["series"]

        if not categories or len(series_names) < 2:
            return self.render_line(chart_data)

        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for i, cat in enumerate(categories):
            row = [cat] + [data_matrix[j][i] if j < len(data_matrix) else 0 for j in range(len(series_names))]
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)
        colors = config.get("colors", []) if config else []

        # ── Primary Line: first series on LEFT Y-axis ──
        line1 = LineChart()
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        ref1 = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        line1.add_data(ref1, titles_from_data=True)
        line1.set_categories(cats_ref)
        line1.y_axis.numFmt = '#,##0'

        # Style first line series
        color1 = colors[0].lstrip("#") if colors else self._color(0)
        line1.series[0].graphicalProperties.line.solidFill = color1
        line1.series[0].graphicalProperties.line.width = 22000
        line1.series[0].smooth = False
        line1.series[0].marker = openpyxl.chart.marker.Marker(symbol="circle", size=5)
        line1.series[0].marker.graphicalProperties.solidFill = color1

        # ── Secondary Line: remaining series on RIGHT Y-axis ──
        line2 = LineChart()
        for idx in range(1, len(series_names)):
            ref = Reference(self._data_ws, min_col=2 + idx, min_row=start, max_row=end)
            line2.add_data(ref, titles_from_data=True)

        line2.y_axis.axId = 200
        line2.y_axis.numFmt = '#,##0'
        line2.y_axis.crosses = "max"

        # Per-chart right axis scale
        axis_cfg = self.RIGHT_AXIS_CONFIG.get(chart_id)
        if axis_cfg is not None:
            rmin, rmax, rstep, rfmt = axis_cfg
            line2.y_axis.scaling.min = rmin
            line2.y_axis.scaling.max = rmax
            line2.y_axis.scaling.majorUnit = rstep
            line2.y_axis.numFmt = rfmt

        # Style second line series
        for idx, series in enumerate(line2.series):
            s_idx = idx + 1
            color = colors[s_idx].lstrip("#") if s_idx < len(colors) else self._color(s_idx)
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 22000
            series.smooth = False
            series.marker = openpyxl.chart.marker.Marker(symbol="circle", size=5)
            series.marker.graphicalProperties.solidFill = color

        line1 += line2

        left_color = colors[0].lstrip("#") if colors else None
        right_color = colors[1].lstrip("#") if len(colors) > 1 else None
        self._apply_style(line1, chart_data.get("title", "Dual-Line Chart"),
                          left_axis_color=left_color, right_axis_color=right_color)
        return line1

    def render_radar(self, chart_data: Dict) -> Optional[RadarChart]:
        """Radar chart.

        Data format:
          - items: list of {label: dimension_name, value: max_scale}
          - config.series: list of person/series names
          - config.seriesData: dict[person_name, list of {label: dimension_name, value: score}]
        """
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})
        if not items:
            return None

        series_names = config.get("series", []) if config else []
        series_data = config.get("seriesData", {}) if config else {}
        if not series_names:
            return None

        # Build category order from items (dimension labels)
        categories = [item.get("label", "") for item in items]

        # Build lookup: dimension -> index for aligning series data
        cat_index = {cat: i for i, cat in enumerate(categories)}

        headers = [_t("Category", "类别", self.language)] + series_names
        rows = []
        for cat in categories:
            row = [cat]
            for person in series_names:
                person_data = series_data.get(person, [])
                val = 0.0
                for datum in person_data:
                    if datum.get("label") == cat:
                        val = float(datum.get("value", 0))
                        break
                row.append(val)
            rows.append(row)

        sheet_title, start, end = self._write_data(headers, rows)

        chart = RadarChart()
        chart.type = "filled"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx in range(2, 2 + len(series_names)):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        colors = config.get("colors", []) if config else []
        for i, series in enumerate(chart.series):
            color = colors[i].lstrip("#") if i < len(colors) else self._color(i)
            # Fill: use original color; alpha (12%) is injected via post-processing
            series.graphicalProperties.solidFill = color
            # Line: original solid color, width matching web strokeWidth=2
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 18000  # ~2pt in EMUs
            # Add markers so data points are visible even when fills overlap
            series.marker = openpyxl.chart.marker.Marker(symbol="circle", size=4)
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.noFill = True

        self._apply_style(chart, chart_data.get("title", "Radar Chart"))
        return chart

    def render_heatmap(self, ws, chart_data: Dict, start_row: int) -> int:
        """Heatmap using conditional formatting. Returns next row.

        Supports two formats:
          - weekday heatmap (cross-change-heatmap): label = "dow|hour|changeCount|incidentCount"
          - fulfillment heatmap (wf-efficiency-heatmap): label = "person|category|displayLabel", value = avgHours
        """
        items = chart_data.get("items", [])
        if not items:
            return start_row

        chart_id = chart_data.get("id", "")
        config = chart_data.get("config", {})

        if chart_id == "cross-change-heatmap":
            return self._render_weekday_heatmap(ws, chart_data, start_row)
        return self._render_fulfillment_heatmap(ws, chart_data, start_row)

    def _render_weekday_heatmap(self, ws, chart_data: Dict, start_row: int) -> int:
        """Weekday x Hour heatmap matching the web page."""
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})

        # Parse label = "dow|hour|changeCount|incidentCount"
        cells: Dict[Tuple[int, int], Dict[str, float]] = {}
        max_change = 0.0
        for item in items:
            label = str(item.get("label", ""))
            parts = label.split("|")
            if len(parts) < 4:
                continue
            try:
                dow = int(parts[0])
                hour = int(parts[1])
                change_count = float(parts[2]) if parts[2] else 0.0
                incident_count = float(parts[3]) if parts[3] else 0.0
            except (ValueError, TypeError):
                continue
            cells[(dow, hour)] = {"change": change_count, "incident": incident_count}
            if change_count > max_change:
                max_change = change_count

        if not cells:
            return start_row

        lang = self.language
        weekday_labels = [
            _t("Mon", "周一", lang),
            _t("Tue", "周二", lang),
            _t("Wed", "周三", lang),
            _t("Thu", "周四", lang),
            _t("Fri", "周五", lang),
            _t("Sat", "周六", lang),
            _t("Sun", "周日", lang),
        ]
        x_axis_label = config.get("xAxisLabel") or _t("Hour", "时段", lang)
        y_axis_label = config.get("yAxisLabel") or _t("Weekday", "星期", lang)
        change_density_label = _t("Change Density", "变更密度", lang)
        incident_hotspots_label = _t("Incident Hotspots", "事件热点", lang)

        title = chart_data.get("title", _t("Heatmap", "热力图", lang))
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12, color=PRIMARY)
        start_row += 2

        # Y-axis label (top-left corner)
        ws.cell(row=start_row, column=1, value=y_axis_label)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="right", vertical="center")

        # X-axis label row (hours 0-23)
        hours = list(range(24))
        for j, hour in enumerate(hours, 2):
            cell = ws.cell(row=start_row, column=j, value=f"{hour}h")
            cell.font = Font(bold=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row += 1

        # Data rows: dow 1-7 (Mon-Sun)
        data_start_row = start_row

        for dow in range(1, 8):
            ws.cell(row=start_row, column=1, value=weekday_labels[dow - 1])
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
            for hour in hours:
                cell = ws.cell(row=start_row, column=2 + hour, value=0)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)
                info = cells.get((dow, hour))
                if info:
                    val = int(info["change"]) if info["change"] == int(info["change"]) else info["change"]
                    cell.value = val
                    if info["incident"] > 0:
                        # Use Excel comment indicator (red triangle) instead of border
                        # to avoid cell deformation across different Excel clients
                        cell.comment = Comment(incident_hotspots_label, "BI Dashboard")
            start_row += 1
        data_end_row = start_row - 1

        # Conditional formatting: blue opacity scale matching web
        end_col = get_column_letter(1 + len(hours))
        rule = ColorScaleRule(
            start_type="min", start_color="E8F0F7",
            end_type="max", end_color="5B8DB8",
        )
        ws.conditional_formatting.add(
            f"B{data_start_row}:{end_col}{data_end_row}", rule
        )

        # Column widths — mark heatmap columns so tables don't override them
        ws.column_dimensions["A"].width = 12
        for j in range(2, 2 + len(hours)):
            col_dim = ws.column_dimensions[get_column_letter(j)]
            col_dim.width = 4.5
            col_dim._heatmap_width = True

        # Legend at bottom
        start_row += 1
        legend_row = start_row
        ws.cell(row=legend_row, column=1, value=change_density_label)
        ws.cell(row=legend_row, column=1).font = Font(bold=True, size=9)

        # Color scale mini bar in cells
        for idx, color in enumerate(["E8F0F7", "B8D4E8", "88B8D9", "5B8DB8"], 2):
            cell = ws.cell(row=legend_row, column=idx, value="")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.cell(row=legend_row, column=6, value=f"▲ {incident_hotspots_label}")
        ws.cell(row=legend_row, column=6).font = Font(size=9, color="EF4444", bold=True)

        return legend_row + 2

    def _render_fulfillment_heatmap(self, ws, chart_data: Dict, start_row: int) -> int:
        """Fulfillment heatmap: rows=people, cols=categories, cells=avgHours.

        Replicates the web's 4-color threshold scale:
          <= 12h  → green   (opacity 0.35–0.65)
          12–20h  → yellow  (opacity 0.55)
          20–28h  → orange  (opacity 0.55)
          >  28h  → red     (opacity 0.55–0.85)
        """
        items = chart_data.get("items", [])
        config = chart_data.get("config", {})

        y_order: List[str] = []
        x_order: List[str] = []
        y_seen = set()
        x_seen = set()
        cell_map: Dict[Tuple[str, str], Dict[str, Any]] = {}

        for item in items:
            label = str(item.get("label", ""))
            parts = label.split("|")
            if len(parts) < 3:
                continue
            person, category, display_label = parts[0], parts[1], parts[2]
            if person not in y_seen:
                y_order.append(person)
                y_seen.add(person)
            if category not in x_seen:
                x_order.append(category)
                x_seen.add(category)
            raw_value = item.get("value", 0)
            cell_map[(person, category)] = {
                "value": float(raw_value) if raw_value is not None else 0.0,
                "display": display_label,
            }

        if not y_order or not x_order:
            return start_row

        lang = self.language
        x_axis_label = config.get("xAxisLabel") or _t("Category", "类别", lang)
        y_axis_label = config.get("yAxisLabel") or _t("Person", "人员", lang)

        title = chart_data.get("title", _t("Heatmap", "热力图", lang))
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12, color=PRIMARY)
        start_row += 2

        ws.cell(row=start_row, column=1, value=y_axis_label)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        for j, cat in enumerate(x_order, 2):
            cell = ws.cell(row=start_row, column=j, value=_localize(cat, lang))
            cell.font = Font(bold=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        start_row += 1

        data_start_row = start_row
        for person in y_order:
            ws.cell(row=start_row, column=1, value=person)
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
            for xi, cat in enumerate(x_order):
                cell = ws.cell(row=start_row, column=2 + xi, value="")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                info = cell_map.get((person, cat))
                if info:
                    cell.value = info["display"]
                    bg, fg = self._fulfillment_cell_colors(info["value"])
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    cell.font = Font(size=9, color=fg)
                else:
                    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            start_row += 1

        ws.column_dimensions["A"].width = 20
        for j in range(2, 2 + len(x_order)):
            col_dim = ws.column_dimensions[get_column_letter(j)]
            col_dim.width = 12
            col_dim._heatmap_width = True

        return start_row + 2

    def _fulfillment_cell_colors(self, val: float) -> Tuple[str, str]:
        """Return (bg_hex, fg_hex) matching the web's 4-color threshold scale.

        Web uses rgba on a white-ish background. We blend with white to get solid fills.
        Foreground is chosen by luminance (light bg → dark text, dark bg → light text).
        """
        def clamp(v: float, lo: float, hi: float) -> float:
            return max(lo, min(hi, v))

        def lerp(a: int, b: int, t: float) -> int:
            return round(a + (b - a) * t)

        def blend(r: int, g: int, b: int, a: float) -> Tuple[int, int, int]:
            """Blend rgba with white background."""
            return (
                int(r * a + 255 * (1 - a)),
                int(g * a + 255 * (1 - a)),
                int(b * a + 255 * (1 - a)),
            )

        def luminance(r: int, g: int, b: int) -> float:
            return 0.299 * r + 0.587 * g + 0.114 * b

        if val <= 12:
            t = clamp(val / 12, 0, 1)
            opacity = 0.35 + t * 0.3  # 0.35 → 0.65
            rgb = blend(16, 185, 129, opacity)
        elif val <= 20:
            t = clamp((val - 12) / 8, 0, 1)
            r = lerp(16, 245, t)
            g = lerp(185, 158, t)
            b = lerp(129, 11, t)
            rgb = blend(r, g, b, 0.55)
        elif val <= 28:
            t = clamp((val - 20) / 8, 0, 1)
            r = lerp(245, 239, t)
            g = lerp(158, 68, t)
            b = lerp(11, 68, t)
            rgb = blend(r, g, b, 0.55)
        else:
            t = clamp((val - 28) / 20, 0, 1)
            opacity = 0.55 + t * 0.3  # 0.55 → 0.85
            rgb = blend(239, 68, 68, opacity)

        bg = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
        fg = "FFFFFF" if luminance(*rgb) < 140 else "1F2937"
        return bg, fg

    def render_bubble(self, chart_data: Dict) -> None:
        """Bubble chart — not supported natively by openpyxl. Return None (will be skipped)."""
        logger.warning("Bubble chart not supported natively, skipping: %s", chart_data.get("title"))
        return None

    def render_doughnut(self, chart_data: Dict) -> Optional[DoughnutChart]:
        """Doughnut chart — used for gauges."""
        items = chart_data.get("items", [])
        if not items:
            return None

        labels = [str(i.get("label", f"Item {idx}")) for idx, i in enumerate(items)]
        values = [float(v) if (v := i.get("value", 0)) is not None else 0.0 for i in items]
        if sum(values) == 0:
            return None

        sheet_title, start, end = self._write_data(
            [_t("Category", "类别", self.language), _t("Value", "数值", self.language)],
            [[l, v] for l, v in zip(labels, values)]
        )

        chart = DoughnutChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, chart_data.get("title", "Doughnut Chart"), CHART_SMALL_W, CHART_SMALL_H)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        return chart


# =============================================================================
# Main Builder
# =============================================================================

class BiXlsxBuilder:
    """Builds an 8-sheet Excel workbook from BI API snapshot data."""

    def __init__(self, snapshot_data: Dict, language: str = "zh", period_label: str = ""):
        """
        Args:
            snapshot_data: Output from BiDataAdapter.get_all_tabs()
                           Dict[str, Dict] mapping tab_id -> normalized tab content
            language: "zh" or "en"
            period_label: Reporting period label shown in footer (e.g. "2024-03-13 ~ 2024-06-11")
        """
        self.data = snapshot_data
        self.language = language
        self.period_label = period_label
        self.styles = XlsxStyles(language)
        self.wb: Optional[openpyxl.Workbook] = None
        self.charts: Optional[BiChartRenderer] = None

    # ─── Utility methods ─────────────────────────────────────────────────

    def _write_title(self, ws, title: str, subtitle: str = "", row: int = 1) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h1
        cell.alignment = self.styles.align_h1
        ws.row_dimensions[row].height = ROW_HEIGHTS["h1"]
        row += 1

        if subtitle:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            cell = ws.cell(row=row, column=1, value=subtitle)
            cell.font = self.styles.font_h2
            cell.alignment = self.styles.align_left
            ws.row_dimensions[row].height = ROW_HEIGHTS["h2"]
            row += 1

        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_section_title(self, ws, title: str, row: int) -> int:
        """Write a section heading (h3 level) with spacing."""
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h3
        cell.alignment = self.styles.align_left
        row += 1
        return row

    def _write_cards(self, ws, cards: List[Dict], row: int) -> int:
        """Write metric cards in a grid (up to 4 per row)."""
        if not cards:
            return row

        col = 1
        value_row = row
        label_row = row + 1
        for i, card in enumerate(cards):
            if i > 0 and i % 4 == 0:
                value_row += 3
                label_row = value_row + 1
                col = 1

            label = card.get("label", "")
            value = str(card.get("value", ""))
            tone = card.get("tone", "neutral")

            ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row, end_column=col + 1)
            vcell = ws.cell(row=value_row, column=col, value=value)
            vcell.font = self.styles.font_kpi_val
            vcell.alignment = self.styles.align_center
            vcell.fill = self.styles.semantic_fill(tone)
            vcell.border = self.styles.border_kpi

            ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=col + 1)
            lcell = ws.cell(row=label_row, column=col, value=label)
            lcell.font = self.styles.font_kpi_label
            lcell.alignment = self.styles.align_center
            lcell.border = self.styles.border_kpi

            ws.row_dimensions[value_row].height = 36
            ws.row_dimensions[label_row].height = 18
            col += 3

        return label_row + 2

    def _write_table(self, ws, table: Dict, row: int) -> int:
        """Write a data table. Returns next row."""
        columns = table.get("columns", [])
        rows_data = table.get("rows", [])
        if not columns:
            return row

        # Table title
        title = table.get("title", "")
        if title:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = self.styles.font_h3
            cell.alignment = self.styles.align_left
            ws.row_dimensions[row].height = ROW_HEIGHTS["h3"]
            row += 1

        # Header
        for ci, h in enumerate(columns, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = self.styles.font_th
            cell.fill = self.styles.fill_header
            cell.alignment = self.styles.align_center
            cell.border = self.styles.border_header
        ws.row_dimensions[row].height = ROW_HEIGHTS["th"]
        row += 1

        # Data rows
        for ri, data_row in enumerate(rows_data):
            fill = self.styles.fill_zebra if ri % 2 == 0 else self.styles.fill_white
            for ci, val in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = self.styles.font_td
                cell.fill = fill
                cell.border = self.styles.border_row
                if isinstance(val, (int, float)):
                    cell.alignment = self.styles.align_right
                    cell.font = self.styles.font_td_num
                else:
                    cell.alignment = self.styles.align_left
            ws.row_dimensions[row].height = ROW_HEIGHTS["td"]
            row += 1

        # Auto width
        all_rows = [columns] + rows_data
        for ci in range(1, len(columns) + 1):
            letter = get_column_letter(ci)
            col_dim = ws.column_dimensions[letter]
            # Don't override heatmap column widths
            if getattr(col_dim, '_heatmap_width', False):
                continue
            max_len = max(
                (len(str(r[ci - 1])) if ci - 1 < len(r) else 0 for r in all_rows),
                default=8,
            )
            col_dim.width = min(max(max_len + 4, 10), 50)

        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_chart(self, ws, chart_data: Dict, row: int) -> int:
        """Render a chart and add to worksheet. Returns next row."""
        chart_type = chart_data.get("type", "bar")
        items = chart_data.get("items", [])

        # Skip charts with no data (avoid empty boxes)
        if not items and chart_type != "heatmap":
            return row

        # Heatmap is special — writes cells directly, not a chart object
        if chart_type == "heatmap":
            return self.charts.render_heatmap(ws, chart_data, row)

        # Write a section title before the chart
        chart_title = chart_data.get("title", "")
        if chart_title:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=chart_title)
            cell.font = self.styles.font_h3
            row += 3  # extra gap so title doesn't overlap chart top / X-axis labels

        chart_obj = self.charts.render(chart_data)
        if chart_obj is None:
            return row

        anchor = f"A{row}"
        ws.add_chart(chart_obj, anchor)
        rows_skip = CHART_ROWS_SMALL if chart_type in ("doughnut",) else CHART_ROWS_STANDARD
        return row + rows_skip + 4  # +4 padding so chart bottom/legend doesn't overlap next content

    def _write_footer(self, ws, row: int) -> int:
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        period = f" • {self.period_label}" if self.period_label else ""
        label = _t(
            f"Generated by BI Dashboard Export{period} • {ts}",
            f"由 BI 仪表板导出{period} • {ts}",
            self.language,
        )
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = self.styles.font_footnote
        cell.alignment = self.styles.align_center
        return row + 1

    # ─── Sheet builders ──────────────────────────────────────────────────

    def _build_generic_sheet(self, ws, tab_data: Dict) -> None:
        """Generic sheet builder for tabs with cards/charts/tables."""
        row = self._write_title(ws, tab_data.get("label", ""), tab_data.get("description", ""))

        # Cards
        cards = tab_data.get("cards", [])
        if cards:
            row = self._write_cards(ws, cards, row)

        # Charts
        charts = tab_data.get("charts", [])
        for chart in charts:
            row = self._write_chart(ws, chart, row)

        # Tables
        tables = tab_data.get("tables", [])
        for table in tables:
            row = self._write_table(ws, table, row)

        self._write_footer(ws, row)

    def _build_executive_sheet(self, ws, tab_data: Dict) -> None:
        """Sheet 1 — Executive Summary with special hero section.

        Layout matches the web page:
          1. Hero Score
          2. Core KPIs  (incident-sla-rate, incident-mttr, change-success-rate, etc.)
          3. Process Health  (process-incident, process-change, process-request, process-problem)
          4. Trend Chart
          5. Risk Summary Table
        """
        row = self._write_title(ws, tab_data.get("label", ""), tab_data.get("description", ""))

        cards = tab_data.get("cards", [])

        # --- Section: Health Score ---
        hero_card = None
        if cards and cards[0].get("id") == "hero-score":
            hero_card = cards[0]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value=hero_card.get("value", ""))
            cell.font = Font(name="Arial", size=36, bold=True, color=PRIMARY)
            cell.alignment = self.styles.align_center
            ws.row_dimensions[row].height = 50

            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
            grade = hero_card.get("tone", "neutral")
            grade_text = _t("Stable", "稳定", self.language) if grade == "success" else \
                        _t("Watch", "关注", self.language) if grade == "warning" else \
                        _t("Risk", "风险", self.language) if grade == "danger" else ""
            cell = ws.cell(row=row, column=5, value=grade_text)
            cell.font = Font(name="Arial", size=24, bold=True, color=SEMANTIC.get(grade, SEMANTIC["neutral"])[1])
            cell.alignment = self.styles.align_center
            ws.row_dimensions[row].height = 50
            row += 2

        # Categorize remaining cards
        core_kpi_cards = []
        process_health_cards = []
        for card in cards[1:] if hero_card else cards:
            cid = card.get("id", "")
            if cid.startswith("process-"):
                process_health_cards.append(card)
            elif cid.startswith("risk-"):
                # Skip risk count cards — they are shown as badges in web UI,
                # and detailed risks are in the Risk Summary table below.
                continue
            else:
                core_kpi_cards.append(card)

        # --- Section: Core KPIs ---
        if core_kpi_cards:
            row = self._write_section_title(ws,
                _t("Core KPIs", "核心 KPI", self.language), row)
            row = self._write_cards(ws, core_kpi_cards, row)

        # --- Section: Process Health ---
        if process_health_cards:
            row = self._write_section_title(ws,
                _t("Process Health", "流程健康", self.language), row)
            row = self._write_cards(ws, process_health_cards, row)

        # --- Section: Trend Chart ---
        charts = tab_data.get("charts", [])
        for chart in charts:
            row = self._write_chart(ws, chart, row)

        # --- Section: Risk Summary ---
        tables = tab_data.get("tables", [])
        if tables:
            row = self._write_section_title(ws,
                _t("Risk Summary", "风险摘要", self.language), row)
        for table in tables:
            row = self._write_table(ws, table, row)

        self._write_footer(ws, row)

    def _build_sla_sheet(self, ws, tab_data: Dict) -> None:
        """Sheet 2 — SLA Analysis with Incident SLA / Request SLA split layout."""
        row = self._write_title(ws, tab_data.get("label", ""), tab_data.get("description", ""))

        cards = tab_data.get("cards", [])
        charts = tab_data.get("charts", [])
        tables = tab_data.get("tables", [])

        # Split by ID prefix (same logic as BI frontend)
        incident_cards = [c for c in cards if c.get("id", "").startswith("sla-")]
        request_cards = [c for c in cards if c.get("id", "").startswith("req-sla-")]
        incident_charts = [c for c in charts if not c.get("id", "").startswith("req-sla-")]
        request_charts = [c for c in charts if c.get("id", "").startswith("req-sla-")]
        incident_tables = [t for t in tables if t.get("id") == "sla-violation-samples"]
        request_tables = [t for t in tables if t.get("id") == "req-sla-violation-samples"]

        # --- Section: Incident SLA ---
        row = self._write_section_title(ws,
            _t("Incident SLA", "事件 SLA", self.language), row)
        if incident_cards:
            row = self._write_cards(ws, incident_cards, row)
        for chart in incident_charts:
            row = self._write_chart(ws, chart, row)
        for table in incident_tables:
            row = self._write_table(ws, table, row)

        # --- Section: Request SLA ---
        row = self._write_section_title(ws,
            _t("Request SLA", "请求 SLA", self.language), row)
        if request_cards:
            row = self._write_cards(ws, request_cards, row)
        for chart in request_charts:
            row = self._write_chart(ws, chart, row)
        for table in request_tables:
            row = self._write_table(ws, table, row)

        self._write_footer(ws, row)

    # ─── Build & Save ────────────────────────────────────────────────────

    def build(self) -> openpyxl.Workbook:
        """Create the workbook with all 8 sheets."""
        self.wb = openpyxl.Workbook()
        self.charts = BiChartRenderer(self.wb, self.language)
        names = SHEET_NAMES[self.language]

        tab_order = [
            "executive-summary",
            "sla-analysis",
            "incident-analysis",
            "change-analysis",
            "request-analysis",
            "problem-analysis",
            "cross-process",
            "workforce",
        ]

        for i, tab_id in enumerate(tab_order):
            tab_data = self.data.get(tab_id, {"id": tab_id, "label": names[i], "cards": [], "charts": [], "tables": []})

            if i == 0:
                ws = self.wb.active
                ws.title = names[i]
            else:
                ws = self.wb.create_sheet(title=names[i])

            ws.sheet_properties.tabColor = TAB_COLORS[i]

            try:
                if tab_id == "executive-summary":
                    self._build_executive_sheet(ws, tab_data)
                elif tab_id == "sla-analysis":
                    self._build_sla_sheet(ws, tab_data)
                else:
                    self._build_generic_sheet(ws, tab_data)
            except Exception as e:
                logger.error("Failed to build sheet '%s': %s", names[i], e)
                ws.cell(row=1, column=1, value=f"Error building sheet: {e}")

        # Print settings
        for ws in self.wb.worksheets:
            if ws.title.startswith("_"):
                continue
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        return self.wb

    def save(self, output_dir: Path = None, filename: str = None) -> Path:
        """Build workbook and save to file."""
        if self.wb is None:
            self.build()

        if output_dir is None:
            output_dir = Path("output")
        output_dir.mkdir(parents=True, exist_ok=True)

        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = "CN" if self.language == "zh" else "EN"
            filename = f"BI_Dashboard_Export_{ts}_{suffix}.xlsx"

        output_path = output_dir / filename

        # Save to memory first so we can post-process chart XML
        bio = io.BytesIO()
        self.wb.save(bio)
        bio.seek(0)
        processed = self._post_process(bio)
        output_path.write_bytes(processed.getvalue())

        logger.info("BI Dashboard export saved to %s", output_path)
        return output_path

    @staticmethod
    def _post_process(src_bio: io.BytesIO) -> io.BytesIO:
        """Post-process xlsx to inject radar chart fill alpha (12%).

        openpyxl does not support alpha/opacity on chart series fills natively.
        We add <a:alpha val="12000"/> inside each <a:srgbClr> inside
        <c:radarChart>/<c:ser>/<c:spPr> so overlapping filled areas remain
        visible in Excel/WPS.

        Only real chart part files (xl/charts/*.xml) are parsed, and we verify
        the root element namespace before treating the content as chart XML.
        """
        dst_bio = io.BytesIO()
        ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
        ns_c = "http://schemas.openxmlformats.org/drawingml/2006/chart"
        chart_ns = f"{{{ns_c}}}"

        with zipfile.ZipFile(src_bio, "r") as zin:
            with zipfile.ZipFile(dst_bio, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/charts/") and item.filename.endswith(".xml"):
                        try:
                            root = etree.fromstring(data)
                        except etree.LxmlError as e:
                            logger.warning("Skipping unreadable chart XML %s: %s", item.filename, e)
                            zout.writestr(item, data)
                            continue

                        # Only mutate files whose root is a chartSpace in the chart namespace
                        if not root.tag.startswith(chart_ns):
                            zout.writestr(item, data)
                            continue

                        # Inject alpha into radar chart series fills only
                        for radar in root.findall(f".//{{{ns_c}}}radarChart"):
                            for ser in radar.findall(f"{{{ns_c}}}ser"):
                                spPr = ser.find(f"{{{ns_c}}}spPr")
                                if spPr is None:
                                    continue
                                srgb = spPr.find(f".//{{{ns_a}}}srgbClr")
                                if srgb is not None and srgb.find(f"{{{ns_a}}}alpha") is None:
                                    alpha = etree.SubElement(srgb, f"{{{ns_a}}}alpha")
                                    alpha.set("val", "12000")  # 12%

                        data = etree.tostring(
                            root, xml_declaration=True, encoding="UTF-8", standalone=True
                        )
                    zout.writestr(item, data)
        dst_bio.seek(0)
        return dst_bio


# Backwards compatibility
XlsxBuilder = BiXlsxBuilder
